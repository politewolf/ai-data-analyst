from fastapi import UploadFile
from sqlalchemy.orm import Session
from app.schemas.file_schema import FileSchema, FileSchemaWithMetadata
from app.models.file import File
import uuid
from app.models.report import Report
from app.models.user import User
from app.models.organization import Organization
from app.models.sheet_schema import SheetSchema
from typing import Optional
from fastapi import HTTPException
from app.models.file import report_file_association
from app.models.file_tag import FileTag
from sqlalchemy.ext.asyncio import AsyncSession
import aiofiles
from sqlalchemy import select, exists
from app.core.telemetry import telemetry
from app.services.file_preview import generate_file_preview
import logging

logger = logging.getLogger(__name__)

class FileService:
    def __init__(self):
        pass

    async def upload_file(self, db: AsyncSession, file: UploadFile, current_user: User, organization: Organization, report_id: Optional[str] = None) -> FileSchema:
        # Generate a unique filename to prevent overwriting existing files
        unique_filename = f"{uuid.uuid4()}_{file.filename}"
        file_location = f"uploads/files/{unique_filename}"

        # Async file writing
        async with aiofiles.open(file_location, "wb") as buffer:
            content = await file.read()
            await buffer.write(content)

        # Create the database entry
        db_file = File(
            filename=file.filename,
            content_type=file.content_type,
            path=file_location,
            user_id=current_user.id,
            organization_id=organization.id
        )

        db.add(db_file)
        await db.commit()
        await db.refresh(db_file)
        
        # Telemetry: file uploaded (minimal fields only)
        try:
            await telemetry.capture(
                "file_uploaded",
                {
                    "file_id": str(db_file.id),
                    "content_type": db_file.content_type,
                    "bytes": len(content or b""),
                    "report_id": report_id,
                },
                user_id=current_user.id,
                org_id=organization.id,
            )
        except Exception:
            pass
        
        # Associate with report if provided
        if report_id:
            stmt = select(Report).filter(Report.id == report_id)
            result = await db.execute(stmt)
            report = result.scalar_one_or_none()
            
            if report:
                report.files.append(db_file)
                await db.commit()
                await db.refresh(report)

        # Generate raw preview (no LLM) - fast, instant
        try:
            db_file.preview = generate_file_preview(db_file)
            db.add(db_file)
            await db.commit()
            await db.refresh(db_file)
            logger.info(f"Generated preview for file {db_file.filename} (type: {db_file.preview.get('type', 'unknown') if db_file.preview else 'none'})")
        except Exception as e:
            # Preview generation failure is non-fatal - log and continue
            logger.warning(f"Failed to generate preview for {db_file.filename}: {e}")
        
        # Return the file schema
        file_schema = FileSchema.from_orm(db_file)
        
        return file_schema
    
    async def remove_file_from_report(self, db: AsyncSession, file_id: str, report_id: str, organization: Organization, current_user: User):
        stmt = select(Report).filter(Report.id == report_id)
        result = await db.execute(stmt)
        report = result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")

        stmt = select(File).filter(File.id == file_id)
        result = await db.execute(stmt)
        file = result.scalar_one_or_none()
        if not file:
            raise HTTPException(status_code=404, detail="File not found")

        stmt = select(report_file_association).filter_by(
            report_id=report_id, file_id=file_id
        )
        result = await db.execute(stmt)
        association = result.first()
        if not association:
            raise HTTPException(status_code=404, detail="File is not associated with this report")

        await db.execute(
            report_file_association.delete().where(
                (report_file_association.c.report_id == report_id) &
                (report_file_association.c.file_id == file_id)
            )
        )
        await db.commit()

        return True
        
    async def get_files(self, db: AsyncSession, organization: Organization):
        stmt = select(File).filter(File.organization_id == organization.id)
        result = await db.execute(stmt)
        files = result.scalars().all()

        # get files with tags
        for file in files:
            stmt = select(FileTag).filter(FileTag.file_id == file.id)
            result = await db.execute(stmt)
            file.tags = result.scalars().all()

            stmt = select(SheetSchema).filter(SheetSchema.file_id == file.id)
            result = await db.execute(stmt)
            file.schemas = result.scalars().all()

        return files

    async def get_files_by_report(self, db: AsyncSession, report_id: str, organization: Organization):
        stmt = select(Report).filter(Report.id == report_id)
        result = await db.execute(stmt)
        report = result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
            
        files = report.files
        return [FileSchema.from_orm(file) for file in files]

    # ==========================================================================
    # DEPRECATED: LLM-based schema extraction methods
    # These methods are no longer called during file upload.
    # We now use raw preview generation instead (see generate_file_preview).
    # Kept for backward compatibility and potential manual re-processing.
    # ==========================================================================
    
    async def _create_sheet_schemas_legacy(self, db: AsyncSession, file: File, model):
        """
        DEPRECATED: LLM-based Excel schema extraction.
        
        This method uses LLM to extract structured schema from Excel files.
        It is no longer called during upload - we now use raw previews instead.
        Kept for backward compatibility if manual schema extraction is needed.
        """
        import warnings
        warnings.warn(
            "_create_sheet_schemas_legacy is deprecated. Use generate_file_preview instead.",
            DeprecationWarning,
            stacklevel=2
        )
        
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        import xlrd
        from app.ai.agents.excel import ExcelAgent
        
        sheet_names = []
        workbook = None

        if file.content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            try:
                workbook = load_workbook(filename=file.path, read_only=True)
                sheet_names = workbook.sheetnames
            except InvalidFileException as e:
                raise HTTPException(status_code=400, detail=f"Failed to process .xlsx file: {e}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error processing Excel file: {e}")
        
        elif file.content_type == "application/vnd.ms-excel":
            try:
                workbook = xlrd.open_workbook(filename=file.path)
                sheet_names = workbook.sheet_names()
            except xlrd.XLRDError as e:
                raise HTTPException(status_code=400, detail=f"Failed to process .xls file: {e}")
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"Error processing Excel file: {e}")

        if not sheet_names:
            return 0

        try:
            processed_sheets_count = 0
            for index, sheet_name in enumerate(sheet_names):
                ea = ExcelAgent(file, model) 
                schema = ea.get_schema(index)

                if schema and "sheet_name" in schema:
                    sc = SheetSchema(
                        sheet_name=schema["sheet_name"],
                        sheet_index=index,
                        schema=schema,
                        file_id=file.id
                    )
                    db.add(sc)
                    processed_sheets_count += 1

            if processed_sheets_count > 0:
                await db.commit()
            return processed_sheets_count
        
        except Exception as e:
            await db.rollback()
            raise HTTPException(status_code=500, detail=f"Error processing sheet schemas: {e}")

        finally:
            if hasattr(workbook, 'close') and callable(workbook.close):
                workbook.close()

    async def _process_pdf_legacy(self, db: AsyncSession, file: File, model):
        """
        DEPRECATED: LLM-based PDF tag extraction.
        
        This method uses LLM to extract semantic tags from PDF files.
        It is no longer called during upload - we now use raw text preview instead.
        Kept for backward compatibility if manual tag extraction is needed.
        """
        import warnings
        warnings.warn(
            "_process_pdf_legacy is deprecated. Use generate_file_preview instead.",
            DeprecationWarning,
            stacklevel=2
        )
        
        import tiktoken
        from app.ai.agents.doc.doc import DocAgent
        
        da = DocAgent(file, model)
        content = da.get_content()

        tags = []   
        tokenizer = tiktoken.get_encoding("cl100k_base")

        tokens = tokenizer.encode(content)
        chunk_size = 100000
        overlap = 300

        for i in range(0, len(tokens), chunk_size - overlap):
            chunk = tokenizer.decode(tokens[i:i+chunk_size])
            new_tags = da.get_tags_from_text(chunk, tags)
            tags.extend(new_tags)
        
        file_tags = []
        
        for tag in tags:
            file_tag = FileTag(
                key=tag["tag"],
                value=tag["value"],
                file_id=file.id
            )
            file_tags.append(file_tag)
        
        for file_tag in file_tags:
            db.add(file_tag)
        await db.commit()
        
        return tags


    async def create_or_get_report_file_association(self, db: AsyncSession, report_id: str, file_id: str):
        # 1. Fetch Report and File
        report_stmt = select(Report).where(Report.id == report_id)
        report_result = await db.execute(report_stmt)
        report = report_result.scalar_one_or_none()
        if not report:
            raise HTTPException(status_code=404, detail=f"Report not found: {report_id}")

        file_stmt = select(File).where(File.id == file_id)
        file_result = await db.execute(file_stmt)
        file = file_result.scalar_one_or_none()
        if not file:
            raise HTTPException(status_code=404, detail=f"File not found: {file_id}")

        # 2. Check if association already exists (more efficient check)
        # Assuming 'files' is the relationship attribute on the Report model
        # Adjust if the relationship is defined differently
        association_exists_stmt = select(exists().where(
            report_file_association.c.report_id == report_id,
            report_file_association.c.file_id == file_id
        ))
        association_exists = await db.scalar(association_exists_stmt)

        # 3. If not associated, create the association by appending
        if not association_exists:
            try:
                # Append the file to the report's collection. SQLAlchemy handles the insert.
                # Ensure the relationship is correctly defined in your models
                # (e.g., on Report: files = relationship("File", secondary=report_file_association, backref="reports"))
                # If the relationship is defined on the File model instead (e.g., file.reports.append(report)), use that.
                report.files.append(file) 
                db.add(report) # Add the modified report to the session if needed
                await db.commit()
                await db.refresh(report) # Refresh report to potentially load the updated relationship
                print(f"Association created between Report {report_id} and File {file_id}")
                return True # Indicate association was created
            except Exception as e:
                await db.rollback()
                print(f"Error creating association: {e}") # Log the specific error
                # Consider raising a specific exception or HTTPException
                raise HTTPException(status_code=500, detail=f"Failed to create association: {e}")
        else:
            print(f"Association already exists between Report {report_id} and File {file_id}")
            return False # Indicate association already existed
        